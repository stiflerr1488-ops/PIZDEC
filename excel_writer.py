from __future__ import annotations

import logging
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook


LOGGER = logging.getLogger(__name__)


class ExcelWriter:
    headers = [
        "Название",
        "Номер",
        "Галочка (синяя/зеленая/пусто)",
        "хорошее место",
        "Рейтинг",
        "Количество оценок",
        "ВК",
        "ТГ",
        "Ватсап",
        "сайт организации",
        "ссылка на карточку",
    ]

    def __init__(self, full_path: Path, potential_path: Path, flush_every: int = 10) -> None:
        self.full_path = full_path
        self.potential_path = potential_path
        self.flush_every = flush_every
        self.full_workbook = Workbook()
        self.full_sheet = self.full_workbook.active
        self.full_sheet.title = "FULL"
        self.full_sheet.append(self.headers)
        self.potential_workbook = Workbook()
        self.potential_sheet = self.potential_workbook.active
        self.potential_sheet.title = "POTENTIAL"
        self.potential_sheet.append(self.headers)
        self._counter = 0
        self.flush()

    def _set_link_cell(self, sheet, row: int, column: int, text: str, url: str) -> None:
        if not url:
            sheet.cell(row=row, column=column, value="")
            return
        display_text = text or url
        cell = sheet.cell(row=row, column=column, value=display_text)
        cell.hyperlink = url
        cell.style = "Hyperlink"

    def _append_to_sheet(self, sheet, organization: "Organization") -> None:
        data = asdict(organization)
        name = data.get("name", "")
        card_url = data.get("card_url", "")
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=name)
        if card_url:
            name_cell = sheet.cell(row=row, column=1)
            name_cell.hyperlink = card_url
            name_cell.style = "Hyperlink"
        sheet.cell(row=row, column=2, value=data.get("phone", ""))
        sheet.cell(row=row, column=3, value=data.get("verified", ""))
        sheet.cell(row=row, column=4, value=data.get("award", ""))
        sheet.cell(row=row, column=5, value=data.get("rating", ""))
        sheet.cell(row=row, column=6, value=data.get("rating_count", ""))
        self._set_link_cell(sheet, row, 7, name, data.get("vk", ""))
        self._set_link_cell(sheet, row, 8, name, data.get("telegram", ""))
        self._set_link_cell(sheet, row, 9, name, data.get("whatsapp", ""))
        self._set_link_cell(sheet, row, 10, name, data.get("website", ""))
        self._set_link_cell(sheet, row, 11, name, card_url)

    def append(self, organization: "Organization", include_in_potential: bool = True) -> None:
        self._append_to_sheet(self.full_sheet, organization)
        if include_in_potential:
            self._append_to_sheet(self.potential_sheet, organization)
        self._counter += 1
        if self._counter % self.flush_every == 0:
            self.flush()

    def append_many(self, organizations: Iterable["Organization"]) -> None:
        for organization in organizations:
            self.append(organization)

    def flush(self) -> None:
        self.full_path.parent.mkdir(parents=True, exist_ok=True)
        self.potential_path.parent.mkdir(parents=True, exist_ok=True)
        self.full_workbook.save(self.full_path)
        self.potential_workbook.save(self.potential_path)
        LOGGER.info("Saved %s", self.full_path)
        LOGGER.info("Saved %s", self.potential_path)

    def close(self) -> None:
        self.flush()
        self.full_workbook.close()
        self.potential_workbook.close()


from pacser_maps import Organization  # noqa: E402
