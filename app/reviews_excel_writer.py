from __future__ import annotations

import logging
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

from app.reviews_parser import Review


LOGGER = logging.getLogger(__name__)


class ReviewsExcelWriter:
    headers = [
        "Имя пользователя",
        "Оценка",
        "Дата отзыва",
        "Полный текст отзыва",
        "Дата ответа организации",
        "Текст ответа организации",
        "ВСЯ ИНФА",
    ]

    def __init__(self, path: Path, flush_every: int = 10) -> None:
        self.path = path
        self.flush_every = flush_every
        self.workbook = Workbook()
        self.full_sheet = self.workbook.active
        self.full_sheet.title = "FULL"
        self.full_sheet.append(self.headers)
        rating_titles = {
            1: "1 звезда",
            2: "2 звезды",
            3: "3 звезды",
            4: "4 звезды",
            5: "5 звёзд",
        }
        self.rating_sheets = {
            rating: self.workbook.create_sheet(title)
            for rating, title in rating_titles.items()
        }
        for sheet in self.rating_sheets.values():
            sheet.append(self.headers)
        self._counter = 0
        self.flush()

    def _full_info(self, data: dict) -> str:
        parts = [
            data.get("user_name", ""),
            data.get("review_date", ""),
            data.get("review_text", ""),
            data.get("response_date", ""),
            data.get("response_text", ""),
        ]
        return " - ".join(str(part or "") for part in parts)

    def _append_row(self, sheet, data: dict) -> None:
        row = sheet.max_row + 1
        name_cell = sheet.cell(row=row, column=1, value=data.get("user_name", ""))
        profile_url = data.get("user_profile_url", "")
        if profile_url:
            name_cell.hyperlink = profile_url
            name_cell.style = "Hyperlink"
        sheet.cell(row=row, column=2, value=data.get("rating", ""))
        sheet.cell(row=row, column=3, value=data.get("review_date", ""))
        sheet.cell(row=row, column=4, value=data.get("review_text", ""))
        sheet.cell(row=row, column=5, value=data.get("response_date", ""))
        sheet.cell(row=row, column=6, value=data.get("response_text", ""))
        sheet.cell(row=row, column=7, value=self._full_info(data))

    def append(self, review: Review) -> None:
        data = asdict(review)
        self._append_row(self.full_sheet, data)
        rating = data.get("rating", 0)
        if isinstance(rating, int) and rating in self.rating_sheets:
            self._append_row(self.rating_sheets[rating], data)
        self._counter += 1
        if self._counter % self.flush_every == 0:
            self.flush()

    def append_many(self, reviews: Iterable[Review]) -> None:
        for review in reviews:
            self.append(review)

    def flush(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.path)
        LOGGER.info("Сохранил файл: %s", self.path)

    def close(self) -> None:
        self.flush()
        self.workbook.close()
