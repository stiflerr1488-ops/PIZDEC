from __future__ import annotations

import logging
import re
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook


LOGGER = logging.getLogger(__name__)


class ExcelWriter:
    headers = [
        "Название",
        "Номер",
        "Галочка (синяя/зеленая/пусто)",
        "хорошее место",
        "Рейтинг",
        "Количество оценок",
        "ВК",
        "ТГ",
        "Ватсап",
        "сайт организации",
        "ссылка на карточку",
    ]

    def __init__(self, path: Path, flush_every: int = 10) -> None:
        self.path = path
        self.flush_every = flush_every
        self.workbook = Workbook()
        self.full_sheet = self.workbook.active
        self.full_sheet.title = "FULL"
        self.full_sheet.append(self.headers)
        self.potential_sheet = self.workbook.create_sheet("POTENTIAL")
        self.potential_sheet.append(self.headers)
        self._counter = 0
        self.flush()

    def _set_link_cell(self, sheet, row: int, column: int, text: str, url: str) -> None:
        if not url:
            sheet.cell(row=row, column=column, value="")
            return
        display_text = text or url
        cell = sheet.cell(row=row, column=column, value=display_text)
        cell.hyperlink = url
        cell.style = "Hyperlink"

    def _extract_links(self, raw: str) -> list[str]:
        if not raw:
            return []
        matches = re.findall(r"(https?://[^\s,;|]+|www\.[^\s,;|]+)", raw, re.IGNORECASE)
        if matches:
            return [match.strip() for match in matches if match.strip()]
        parts = re.split(r"[\s,;|]+", raw)
        return [part.strip() for part in parts if part.strip() and "." in part]

    def _redistribute_links(
        self,
        *,
        website: str,
        vk: str,
        telegram: str,
        whatsapp: str,
    ) -> tuple[str, str, str, str]:
        links = self._extract_links(website)
        remaining_sites: list[str] = []
        for link in links:
            lower_link = link.lower()
            if not vk and ("vk.com" in lower_link or "vkontakte.ru" in lower_link):
                vk = link
                continue
            if not telegram and ("t.me" in lower_link or "telegram.me" in lower_link):
                telegram = link
                continue
            if not whatsapp and (
                "wa.me" in lower_link
                or "api.whatsapp.com" in lower_link
                or "whatsapp.com" in lower_link
            ):
                whatsapp = link
                continue
            remaining_sites.append(link)
        if links:
            website = remaining_sites[0] if remaining_sites else ""
        return website, vk, telegram, whatsapp

    def _append_to_sheet(self, sheet, organization: "Organization") -> None:
        data = asdict(organization)
        name = data.get("name", "")
        card_url = data.get("card_url", "")
        website, vk, telegram, whatsapp = self._redistribute_links(
            website=data.get("website", ""),
            vk=data.get("vk", ""),
            telegram=data.get("telegram", ""),
            whatsapp=data.get("whatsapp", ""),
        )
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=name)
        if card_url:
            name_cell = sheet.cell(row=row, column=1)
            name_cell.hyperlink = card_url
            name_cell.style = "Hyperlink"
        sheet.cell(row=row, column=2, value=data.get("phone", ""))
        sheet.cell(row=row, column=3, value=data.get("verified", ""))
        sheet.cell(row=row, column=4, value=data.get("award", ""))
        sheet.cell(row=row, column=5, value=data.get("rating", ""))
        sheet.cell(row=row, column=6, value=data.get("rating_count", ""))
        self._set_link_cell(sheet, row, 7, "вк", vk)
        self._set_link_cell(sheet, row, 8, "тг", telegram)
        self._set_link_cell(sheet, row, 9, "ватсап", whatsapp)
        self._set_link_cell(sheet, row, 10, "сайт", website)
        self._set_link_cell(sheet, row, 11, "карточка", card_url)

    def append(self, organization: "Organization", include_in_potential: bool = True) -> None:
        self._append_to_sheet(self.full_sheet, organization)
        if include_in_potential:
            self._append_to_sheet(self.potential_sheet, organization)
        self._counter += 1
        if self._counter % self.flush_every == 0:
            self.flush()

    def append_many(self, organizations: Iterable["Organization"]) -> None:
        for organization in organizations:
            self.append(organization)

    def flush(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.path)
        LOGGER.info("Сохранил файл: %s", self.path)

    def close(self) -> None:
        self.flush()
        self.workbook.close()


from app.pacser_maps import Organization  # noqa: E402
